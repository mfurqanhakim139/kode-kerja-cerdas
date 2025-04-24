import openpyxl
from pathlib import Path

# Persiapan: Buat file Excel contoh jika belum ada
nama_file_excel = 'contoh_data.xlsx'
file_excel = Path(nama_file_excel)
if not file_excel.exists():
    try:
        wb_baru = openpyxl.Workbook()
        sheet = wb_baru.active
        sheet.title = 'Data Produk'
        # Buat header
        sheet['A1'] = 'ID Produk'
        sheet['B1'] = 'Nama Produk'
        sheet['C1'] = 'Harga'
        # Isi data contoh
        data = [
            ('P001', 'Laptop ABC', 12000000),
            ('P002', 'Keyboard Mekanik', 850000),
            ('P003', 'Mouse Gaming', 450000)
        ]
        for row_idx, row_data in enumerate(data, start=2): # Mulai dari baris 2
            sheet.cell(row=row_idx, column=1, value=row_data[0])
            sheet.cell(row=row_idx, column=2, value=row_data[1])
            sheet.cell(row=row_idx, column=3, value=row_data[2])
        wb_baru.save(nama_file_excel)
        print(f"File '{nama_file_excel}' berhasil dibuat untuk demo.")
    except Exception as e:
        print(f"Gagal membuat file Excel demo: {e}")


# --- Membaca File Excel ---
try:
    # 1. Muat Workbook
    workbook = openpyxl.load_workbook(nama_file_excel)
    print(f"Berhasil memuat: {nama_file_excel}")

    # 2. Dapatkan Sheet
    print(f"Nama-nama sheet: {workbook.sheetnames}")
    sheet_produk = workbook['Data Produk'] # Akses berdasarkan nama
    # sheet_aktif = workbook.active # Atau dapatkan sheet aktif

    print(f"Judul sheet aktif: {sheet_produk.title}")

    # 3. Akses Cell spesifik
    sel_a1 = sheet_produk['A1']
    sel_c2 = sheet_produk.cell(row=2, column=3) # Sel C2 (Harga Laptop)

    # 4. Dapatkan Nilai Cell
    print(f"Nilai sel A1: {sel_a1.value}")
    print(f"Nilai sel C2: {sel_c2.value}")
    print(f"Koordinat sel C2: {sel_c2.coordinate}, Baris: {sel_c2.row}, Kolom: {sel_c2.column}")

    # 5. Iterasi data (misal, semua produk dan harganya)
    print("\nData Produk:")
    # iter_rows: min_row=2 untuk skip header
    # min_col=2, max_col=3 untuk ambil kolom B (Nama) dan C (Harga) saja
    for row_obj in sheet_produk.iter_rows(min_row=2, min_col=2, max_col=3):
        nama_produk = row_obj[0].value # Kolom pertama dalam iterasi ini (B)
        harga = row_obj[1].value       # Kolom kedua dalam iterasi ini (C)
        print(f"- Nama: {nama_produk}, Harga: Rp {harga:,.0f}")

except FileNotFoundError:
    print(f"Error: File '{nama_file_excel}' tidak ditemukan.")
except Exception as e:
    print(f"Terjadi error saat membaca Excel: {e}")

