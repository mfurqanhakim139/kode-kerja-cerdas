import openpyxl

nama_file_baru = 'laporan_baru.xlsx'

try:
    # 1. Buat Workbook baru
    wb = openpyxl.Workbook()

    # 2. Dapatkan sheet aktif dan ganti nama
    sheet = wb.active
    sheet.title = 'Ringkasan Penjualan'

    # 4. Menulis ke cell
    sheet['A1'] = 'Bulan'
    sheet['B1'] = 'Pendapatan'
    sheet.cell(row=2, column=1, value='Januari')
    sheet.cell(row=2, column=2, value=15000000)
    sheet.cell(row=3, column=1, value='Februari')
    sheet.cell(row=3, column=2, value=17500000)
    sheet['A4'] = 'Maret'
    sheet['B4'] = 21000000

    # 2. Buat sheet baru
    sheet_detail = wb.create_sheet(title='Detail Transaksi')
    sheet_detail['A1'] = 'ID Transaksi'
    sheet_detail['B1'] = 'Nilai'
    # ... (isi data detail jika perlu)

    # 5. Simpan Workbook
    wb.save(nama_file_baru)
    print(f"Berhasil membuat dan menyimpan workbook ke '{nama_file_baru}'")

except Exception as e:
    print(f"Gagal menulis atau menyimpan file Excel: {e}")
