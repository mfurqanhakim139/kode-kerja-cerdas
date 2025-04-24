import openpyxl
# Impor komponen styling yang diperlukan
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from pathlib import Path # Impor Path untuk konsistensi

nama_file_baru = 'laporan_penjualan_styled.xlsx'

try:
    # 1. Buat Workbook baru
    wb = openpyxl.Workbook()

    # --- Style Definitions ---
    font_header = Font(name='Calibri', size=12, bold=True, color='FF000000') # Font tebal hitam
    font_total = Font(name='Calibri', size=11, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    border_thin_top = Border(top=Side(style='thin'))
    border_bottom_header = Border(bottom=Side(style='thin')) # Garis bawah untuk header
    fill_header = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid') # Abu-abu muda
    # Format Angka Rupiah (Accounting Style)
    format_rp = numbers.FORMAT_NUMBER_COMMA_SEPARATED1.replace('#,##0', '_("Rp"* #,##0_)') + ';[Red]-' + numbers.FORMAT_NUMBER_COMMA_SEPARATED1.replace('#,##0', '_("Rp"* #,##0_)')

    # --- Sheet 1: Ringkasan Penjualan ---
    sheet = wb.active
    sheet.title = 'Ringkasan Penjualan'

    # Menulis header dan menerapkan style
    sheet['A1'] = 'Bulan'
    sheet['B1'] = 'Pendapatan'
    for col in ['A1', 'B1']:
        sheet[col].font = font_header
        sheet[col].alignment = align_center
        sheet[col].fill = fill_header
        sheet[col].border = border_bottom_header

    # Menulis data dan menerapkan style
    data_penjualan = {
        'A2': 'Januari', 'B2': 15000000,
        'A3': 'Februari', 'B3': 17500000,
        'A4': 'Maret', 'B4': 21000000,
    }
    for cell_coord, value in data_penjualan.items():
        sheet[cell_coord] = value
        # Style untuk kolom B (Pendapatan)
        if cell_coord.startswith('B'):
            sheet[cell_coord].number_format = format_rp
            sheet[cell_coord].alignment = align_right

    # Menambah baris Total dan Formula
    total_row = 5
    sheet[f'A{total_row}'] = 'Total'
    sheet[f'B{total_row}'] = f'=SUM(B2:B{total_row-1})' # Formula SUM

    # Menerapkan style pada baris Total
    sheet[f'A{total_row}'].font = font_total
    sheet[f'A{total_row}'].alignment = align_right # Total rata kanan
    sheet[f'B{total_row}'].font = font_total
    sheet[f'B{total_row}'].alignment = align_right
    sheet[f'B{total_row}'].number_format = format_rp
    sheet[f'B{total_row}'].border = border_thin_top # Garis atas untuk total

    # Mengatur lebar kolom agar lebih rapi (opsional)
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20

    # --- Sheet 2: Detail Transaksi ---
    sheet_detail = wb.create_sheet(title='Detail Transaksi')

    # Menulis header sheet 2 dan style
    sheet_detail['A1'] = 'ID Transaksi'
    sheet_detail['B1'] = 'Nilai'
    sheet_detail['C1'] = 'Tanggal' # Contoh tambah kolom
    for col in ['A1', 'B1', 'C1']:
        sheet_detail[col].font = font_header
        sheet_detail[col].alignment = align_center
        sheet_detail[col].fill = fill_header
        sheet_detail[col].border = border_bottom_header
    # Atur lebar kolom sheet 2
    sheet_detail.column_dimensions['A'].width = 15
    sheet_detail.column_dimensions['B'].width = 20
    sheet_detail.column_dimensions['C'].width = 15
    # ... (Anda bisa menambahkan data detail jika perlu) ...


    # 5. Simpan Workbook
    wb.save(nama_file_baru)
    print(f"Berhasil membuat dan menyimpan workbook ke '{nama_file_baru}'")
    print("File berisi data, formula SUM, dan styling dasar.")

except ImportError:
     print("Error: Library 'openpyxl' belum terinstal.")
     print("Silakan instal dengan: pip install openpyxl")
except Exception as e:
    print(f"Gagal menulis atau menyimpan file Excel: {e}")